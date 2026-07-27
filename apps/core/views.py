"""
Vues pour l'application core.
AMÉLIORÉ : Système de permissions RBAC complet, Dashboard adaptatif et Portail unifié.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.template.loader import get_template
from django.utils import timezone
import csv
import os
import io
import calendar
from datetime import datetime, date
from django.core.management import call_command

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

try:
    import qrcode
    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False

from .models import Zone, Caveau, Concession, Defunt, Inhumation, DemandeExhumation, ParametreCimetiere


# ==============================================================================
# DÉCORATEUR DE PERMISSIONS RBAC
# ==============================================================================
def role_permission_required(permission_codename, redirect_url='portal_home'):
    """Décorateur pour vérifier les permissions basées sur le rôle de l'utilisateur."""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('mfa_login')
            if not request.user.has_permission(permission_codename):
                messages.error(request, "Accès refusé : permissions insuffisantes pour cette action.")
                return redirect(redirect_url)
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ==============================================================================
# INITIALISATION DE LA BASE DE DONNÉES
# ==============================================================================
def init_db(request):
    result = []
    try:
        result.append("📦 Application des migrations...")
        call_command('migrate', verbosity=0)
        result.append("✅ Migrations appliquées")
        
        result.append("📁 Collecte des fichiers statiques...")
        call_command('collectstatic', verbosity=0, interactive=False)
        result.append("✅ Fichiers statiques collectés")
        
        from apps.accounts.models import User
        admin_email = os.environ.get('ADMIN_EMAIL', 'betsalimolotha5@gmail.com')
        admin_password = os.environ.get('ADMIN_PASSWORD', '&Andrade2580')
        
        if not User.objects.filter(email=admin_email).exists():
            User.objects.create_superuser(
                email=admin_email,
                password=admin_password,
                first_name='Admin',
                last_name='Système',
            )
            result.append(f"✅ Superutilisateur créé : {admin_email}")
        else:
            result.append(f"⚠️ Superutilisateur existe déjà : {admin_email}")
        
        result.append("<br><br>🎉 <b>INITIALISATION TERMINÉE AVEC SUCCÈS !</b>")
        result.append("<a href='/admin/' style='font-size: 1.2em; color: blue;'>👉 Aller à l'administration</a>")
        
    except Exception as e:
        result.append(f"<br><br>❌ <b>ERREUR :</b> {str(e)}")
    
    html = "<br>".join(result)
    return HttpResponse(f"<h1>Initialisation de la base de données</h1><p>{html}</p>")


# ==============================================================================
# PORTAIL UNIFIÉ (Point d'entrée : Carte + Boutons selon rôle)
# ==============================================================================
@login_required
def portal_home(request):
    """
    Point d'entrée unique du portail : Carte publique + Boutons d'action selon le rôle.
    """
    user = request.user
    context = {
        'user': user,
        'can_manage_caveaux': user.has_permission('create_caveaux'),
        'can_manage_concessions': user.has_permission('create_concessions'),
        'can_manage_inhumations': user.has_permission('create_inhumations'),
        'can_manage_exhumations': user.has_permission('validate_exhumation'),
        'can_view_statistics': user.has_permission('view_statistics'),
        'can_manage_users': user.has_permission('manage_users'),
        'can_manage_settings': user.has_permission('manage_settings'),
    }
    return render(request, 'core/portal_home.html', context)


# ==============================================================================
# DASHBOARD ADAPTATIF (Selon le rôle)
# ==============================================================================
@login_required
def dashboard(request):
    user = request.user
    today = timezone.now().date()
    
    # Stats de base (accessibles à tous les rôles internes)
    total_caveaux = Caveau.objects.count()
    caveaux_occupes = Caveau.objects.filter(statut='OCCUPE').count()
    caveaux_disponibles = Caveau.objects.filter(statut='DISPONIBLE').count()
    taux_occupation = round((caveaux_occupes / total_caveaux * 100), 1) if total_caveaux > 0 else 0
    
    total_concessions = Concession.objects.count()
    concessions_actives = Concession.objects.filter(statut='ACTIVE').count()
    
    context = {
        'total_caveaux': total_caveaux,
        'caveaux_occupes': caveaux_occupes,
        'caveaux_disponibles': caveaux_disponibles,
        'taux_occupation': taux_occupation,
        'total_concessions': total_concessions,
        'concessions_actives': concessions_actives,
    }
    
    if user.is_admin():
        # Dashboard Admin complet + Audit
        labels = []
        data_inhumations = []
        data_revenus = []
        
        for i in range(5, -1, -1):
            month = today.month - i
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            
            _, last_day = calendar.monthrange(year, month)
            first_day = date(year, month, 1)
            end_day = date(year, month, last_day)
            
            labels.append(f"{calendar.month_name[month][:3]} {year}")
            
            inhumations = Inhumation.objects.filter(date_inhumation__range=(first_day, end_day)).count()
            data_inhumations.append(inhumations)
            
            try:
                from apps.billing.models import Paiement
                revenus = Paiement.objects.filter(
                    statut='VALIDE',
                    date_paiement__range=(first_day, end_day)
                ).aggregate(total=Sum('montant'))['total'] or 0
                data_revenus.append(float(revenus))
            except Exception:
                data_revenus.append(0.0)
        
        context.update({
            'role': 'admin',
            'chart_labels': labels,
            'chart_inhumations': data_inhumations,
            'chart_revenus': data_revenus,
            'total_revenus_6_mois': sum(data_revenus),
            'can_view_audit': user.has_permission('view_audit_logs'),
            'can_manage_users': user.has_permission('manage_users'),
            'can_manage_settings': user.has_permission('manage_settings'),
        })
        
    elif user.is_field_agent() or user.is_secretary():
        # Dashboard Opérationnel (Agent / Secrétaire)
        context.update({
            'role': 'staff',
            'can_manage_caveaux': user.has_permission('create_caveaux'),
            'can_validate_reservations': user.has_permission('update_concessions'),
            'recent_inhumations': Inhumation.objects.select_related('defunt', 'concession__caveau').order_by('-date_inhumation')[:10],
        })
        
    elif user.is_client():
        # Dashboard Client
        context.update({
            'role': 'client',
            'mes_concessions': Concession.objects.filter(concessionnaire=user).select_related('caveau', 'caveau__zone', 'defunt'),
            'mes_demandes_exhumation': DemandeExhumation.objects.filter(demandeur=user).order_by('-date_demande'),
        })

    return render(request, 'core/dashboard.html', context)


# ==============================================================================
# EXPORTS CSV (Protégés par permissions RBAC)
# ==============================================================================
@role_permission_required('view_caveaux')
def export_csv_caveaux(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="caveaux_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff'.encode('utf-8'))
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Code', 'Zone', 'Statut', 'Type', 'Longueur (m)', 'Largeur (m)', 'Profondeur (m)', 'Prix concession', 'Prix perpétuité', 'Rangée', 'Numéro place', 'Notes'])
    for caveau in Caveau.objects.select_related('zone').all():
        writer.writerow([caveau.code, caveau.zone.nom if caveau.zone else '', caveau.get_statut_display(), caveau.get_type_caveau_display(), caveau.longueur, caveau.largeur, caveau.profondeur, caveau.prix_concession, caveau.prix_perpetuite, caveau.rangee, caveau.numero_place, caveau.notes])
    return response

@role_permission_required('view_concessions')
def export_csv_concessions(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="concessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff'.encode('utf-8'))
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['N° Contrat', 'Concessionnaire', 'Email', 'Caveau', 'Zone', 'Type', 'Durée', 'Date début', 'Date fin', 'Statut', 'Montant total', 'Montant payé', 'Défunt', 'Notes'])
    for c in Concession.objects.select_related('concessionnaire', 'caveau', 'caveau__zone', 'defunt').all():
        writer.writerow([c.numero_contrat, c.concessionnaire.get_full_name() if c.concessionnaire else '', c.concessionnaire.email if c.concessionnaire else '', c.caveau.code if c.caveau else '', c.caveau.zone.nom if c.caveau and c.caveau.zone else '', c.get_type_concession_display(), c.duree_annees or '', c.date_debut.strftime('%d/%m/%Y') if c.date_debut else '', c.date_fin.strftime('%d/%m/%Y') if c.date_fin else '', c.get_statut_display(), c.montant_total, c.montant_paye, f"{c.defunt.nom} {c.defunt.prenom}" if c.defunt else '', c.notes])
    return response

@role_permission_required('view_defunts')
def export_csv_defunts(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="defunts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff'.encode('utf-8'))
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nom', 'Prénom', 'Date naissance', 'Date décès', 'Lieu décès', 'Sexe', 'N° Identité', 'Nationalité', 'N° Acte décès', 'Caveau', 'Zone', 'Notes'])
    for defunt in Defunt.objects.all():
        caveau, zone = '', ''
        if defunt.concessions.exists():
            concession = defunt.concessions.first()
            if concession.caveau:
                caveau = concession.caveau.code
                zone = concession.caveau.zone.nom if concession.caveau.zone else ''
        writer.writerow([defunt.nom, defunt.prenom, defunt.date_naissance.strftime('%d/%m/%Y') if defunt.date_naissance else '', defunt.date_deces.strftime('%d/%m/%Y') if defunt.date_deces else '', defunt.lieu_deces, defunt.get_sexe_display(), defunt.numero_identite, defunt.nationalite, defunt.numero_acte_deces, caveau, zone, defunt.notes])
    return response

@role_permission_required('view_inhumations')
def export_csv_inhumations(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="inhumations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff'.encode('utf-8'))
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Défunt', 'Caveau', 'Zone', 'Date inhumation', 'Profondeur (m)', 'N° place', 'Notes'])
    for inh in Inhumation.objects.select_related('defunt', 'concession', 'concession__caveau', 'concession__caveau__zone').all():
        writer.writerow([f"{inh.defunt.nom} {inh.defunt.prenom}", inh.concession.caveau.code if inh.concession.caveau else '', inh.concession.caveau.zone.nom if inh.concession.caveau and inh.concession.caveau.zone else '', inh.date_inhumation.strftime('%d/%m/%Y') if inh.date_inhumation else '', inh.profondeur, inh.numero_place_dans_caveau, inh.notes])
    return response

@role_permission_required('view_exhumations')
def export_csv_exhumations(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="exhumations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff'.encode('utf-8'))
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['ID', 'Demandeur', 'Lien parenté', 'Téléphone', 'Défunt', 'Caveau', 'Motif', 'Destination', 'Statut', 'Date demande', 'Date validation', 'Date réalisation', 'Notes'])
    for demande in DemandeExhumation.objects.select_related('inhumation', 'inhumation__defunt', 'inhumation__concession', 'inhumation__concession__caveau').all():
        writer.writerow([demande.id, demande.nom_demandeur, demande.lien_parente, demande.telephone_demandeur, f"{demande.inhumation.defunt.nom} {demande.inhumation.defunt.prenom}" if demande.inhumation.defunt else '', demande.inhumation.concession.caveau.code if demande.inhumation.concession.caveau else '', demande.motif, demande.get_destination_display(), demande.get_statut_display(), demande.date_demande.strftime('%d/%m/%Y %H:%M') if demande.date_demande else '', demande.date_validation.strftime('%d/%m/%Y %H:%M') if demande.date_validation else '', demande.date_realisation.strftime('%d/%m/%Y %H:%M') if demande.date_realisation else '', demande.notes])
    return response


# ==============================================================================
# EXPORTS EXCEL (Protégés par permissions RBAC)
# ==============================================================================
@role_permission_required('view_caveaux')
def export_excel_caveaux(request):
    if not EXCEL_AVAILABLE: return HttpResponse("La bibliothèque openpyxl n'est pas installée.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caveaux"
    ws.append(['Code', 'Zone', 'Statut', 'Type', 'Longueur', 'Largeur', 'Profondeur', 'Prix concession', 'Prix perpétuité', 'Rangée', 'Numéro place', 'Notes'])
    for caveau in Caveau.objects.select_related('zone').all():
        ws.append([caveau.code, caveau.zone.nom if caveau.zone else '', caveau.get_statut_display(), caveau.get_type_caveau_display(), float(caveau.longueur), float(caveau.largeur), float(caveau.profondeur), float(caveau.prix_concession), float(caveau.prix_perpetuite), caveau.rangee, caveau.numero_place, caveau.notes])
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="caveaux_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

@role_permission_required('view_concessions')
def export_excel_concessions(request):
    if not EXCEL_AVAILABLE: return HttpResponse("La bibliothèque openpyxl n'est pas installée.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Concessions"
    ws.append(['N° Contrat', 'Concessionnaire', 'Email', 'Caveau', 'Zone', 'Type', 'Durée', 'Date début', 'Date fin', 'Statut', 'Montant total', 'Montant payé', 'Défunt', 'Notes'])
    for c in Concession.objects.select_related('concessionnaire', 'caveau', 'caveau__zone', 'defunt').all():
        ws.append([c.numero_contrat, c.concessionnaire.get_full_name() if c.concessionnaire else '', c.concessionnaire.email if c.concessionnaire else '', c.caveau.code if c.caveau else '', c.caveau.zone.nom if c.caveau and c.caveau.zone else '', c.get_type_concession_display(), c.duree_annees or '', c.date_debut.strftime('%d/%m/%Y') if c.date_debut else '', c.date_fin.strftime('%d/%m/%Y') if c.date_fin else '', c.get_statut_display(), float(c.montant_total), float(c.montant_paye), f"{c.defunt.nom} {c.defunt.prenom}" if c.defunt else '', c.notes])
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="concessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

@role_permission_required('view_defunts')
def export_excel_defunts(request):
    if not EXCEL_AVAILABLE: return HttpResponse("La bibliothèque openpyxl n'est pas installée.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Défunts"
    ws.append(['Nom', 'Prénom', 'Date naissance', 'Date décès', 'Lieu décès', 'Sexe', 'N° Identité', 'Nationalité', 'N° Acte décès', 'Caveau', 'Zone', 'Notes'])
    for defunt in Defunt.objects.all():
        caveau, zone = '', ''
        if defunt.concessions.exists():
            concession = defunt.concessions.first()
            if concession.caveau:
                caveau = concession.caveau.code
                zone = concession.caveau.zone.nom if concession.caveau.zone else ''
        ws.append([defunt.nom, defunt.prenom, defunt.date_naissance.strftime('%d/%m/%Y') if defunt.date_naissance else '', defunt.date_deces.strftime('%d/%m/%Y') if defunt.date_deces else '', defunt.lieu_deces, defunt.get_sexe_display(), defunt.numero_identite, defunt.nationalite, defunt.numero_acte_deces, caveau, zone, defunt.notes])
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="defunts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

@role_permission_required('view_inhumations')
def export_excel_inhumations(request):
    if not EXCEL_AVAILABLE: return HttpResponse("La bibliothèque openpyxl n'est pas installée.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inhumations"
    ws.append(['Défunt', 'Caveau', 'Zone', 'Date inhumation', 'Profondeur (m)', 'N° place', 'Notes'])
    for inh in Inhumation.objects.select_related('defunt', 'concession', 'concession__caveau', 'concession__caveau__zone').all():
        ws.append([f"{inh.defunt.nom} {inh.defunt.prenom}", inh.concession.caveau.code if inh.concession.caveau else '', inh.concession.caveau.zone.nom if inh.concession.caveau and inh.concession.caveau.zone else '', inh.date_inhumation.strftime('%d/%m/%Y') if inh.date_inhumation else '', float(inh.profondeur), inh.numero_place_dans_caveau, inh.notes])
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="inhumations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

@role_permission_required('view_exhumations')
def export_excel_exhumations(request):
    if not EXCEL_AVAILABLE: return HttpResponse("La bibliothèque openpyxl n'est pas installée.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhumations"
    ws.append(['ID', 'Demandeur', 'Lien parenté', 'Téléphone', 'Défunt', 'Caveau', 'Motif', 'Destination', 'Statut', 'Date demande', 'Date validation', 'Date réalisation', 'Notes'])
    for demande in DemandeExhumation.objects.select_related('inhumation', 'inhumation__defunt', 'inhumation__concession', 'inhumation__concession__caveau').all():
        ws.append([demande.id, demande.nom_demandeur, demande.lien_parente, demande.telephone_demandeur, f"{demande.inhumation.defunt.nom} {demande.inhumation.defunt.prenom}" if demande.inhumation.defunt else '', demande.inhumation.concession.caveau.code if demande.inhumation.concession.caveau else '', demande.motif, demande.get_destination_display(), demande.get_statut_display(), demande.date_demande.strftime('%d/%m/%Y %H:%M') if demande.date_demande else '', demande.date_validation.strftime('%d/%m/%Y %H:%M') if demande.date_validation else '', demande.date_realisation.strftime('%d/%m/%Y %H:%M') if demande.date_realisation else '', demande.notes])
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="exhumations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


# ==============================================================================
# EXPORTS PDF (Protégés par permissions RBAC)
# ==============================================================================
@role_permission_required('generate_contract_pdf')
def contrat_concession_pdf(request, concession_id):
    if not WEASYPRINT_AVAILABLE:
        messages.error(request, "Le système de génération PDF n'est pas disponible.")
        return redirect('portal_home')
    
    user = request.user
    if user.is_admin() or user.is_staff:
        concession = get_object_or_404(Concession, id=concession_id)
    else:
        concession = get_object_or_404(Concession, id=concession_id, concessionnaire=user)
    
    reste_a_payer = max(0, float(concession.montant_total or 0) - float(concession.montant_paye or 0))
    
    context = {
        'concession': concession, 
        'parametres': ParametreCimetiere.objects.first(), 
        'date_generation': timezone.now(), 
        'site_name': 'Gestion Cimetière',
        'reste_a_payer': reste_a_payer
    }
    template = get_template('core/pdf/contrat_concession_pdf.html')
    pdf_file = HTML(string=template.render(context), base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="contrat_{concession.numero_contrat}.pdf"'
    return response

@role_permission_required('generate_attestation_pdf')
def attestation_concession_pdf(request, concession_id):
    if not WEASYPRINT_AVAILABLE:
        messages.error(request, "Le système de génération PDF n'est pas disponible.")
        return redirect('portal_home')
    
    user = request.user
    if user.is_admin() or user.is_staff:
        concession = get_object_or_404(Concession, id=concession_id)
    else:
        concession = get_object_or_404(Concession, id=concession_id, concessionnaire=user)
    
    reste_a_payer = max(0, float(concession.montant_total or 0) - float(concession.montant_paye or 0))
    
    context = {
        'concession': concession, 
        'parametres': ParametreCimetiere.objects.first(), 
        'date_generation': timezone.now(), 
        'site_name': 'Gestion Cimetière',
        'reste_a_payer': reste_a_payer
    }
    template = get_template('core/pdf/attestation_concession_pdf.html')
    pdf_file = HTML(string=template.render(context), base_url=request.build_absolute_uri('/')).write_pdf()
    nom_client = concession.concessionnaire.get_full_name().replace(' ', '_') if concession.concessionnaire.get_full_name() else 'Client'
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Attestation_{nom_client}_{timezone.now().strftime('%Y%m%d')}.pdf"'
    return response

@role_permission_required('generate_pv_inhumation_pdf')
def pv_inhumation_pdf(request, inhumation_id):
    if not WEASYPRINT_AVAILABLE:
        messages.error(request, "Le système de génération PDF n'est pas disponible.")
        return redirect('portal_home')
        
    inhumation = get_object_or_404(Inhumation, id=inhumation_id)
    context = {'inhumation': inhumation, 'parametres': ParametreCimetiere.objects.first(), 'date_generation': timezone.now(), 'site_name': 'Gestion Cimetière'}
    template = get_template('core/pdf/pv_inhumation_pdf.html')
    pdf_file = HTML(string=template.render(context), base_url=request.build_absolute_uri('/')).write_pdf()
    nom_defunt = inhumation.defunt.nom.replace(' ', '_') if inhumation.defunt else 'Defunt'
    date_str = inhumation.date_inhumation.strftime('%Y%m%d') if inhumation.date_inhumation else timezone.now().strftime('%Y%m%d')
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="PV_inhumation_{nom_defunt}_{date_str}.pdf"'
    return response

@role_permission_required('generate_exhumation_pdf')
def autorisation_exhumation_pdf(request, demande_id):
    if not WEASYPRINT_AVAILABLE:
        messages.error(request, "Le système de génération PDF n'est pas disponible.")
        return redirect('portal_home')
        
    demande = get_object_or_404(DemandeExhumation, id=demande_id)
    context = {'demande': demande, 'parametres': ParametreCimetiere.objects.first(), 'date_generation': timezone.now(), 'site_name': 'Gestion Cimetière'}
    template = get_template('core/pdf/autorisation_exhumation_pdf.html')
    pdf_file = HTML(string=template.render(context), base_url=request.build_absolute_uri('/')).write_pdf()
    nom_defunt = demande.inhumation.defunt.nom.replace(' ', '_') if demande.inhumation and demande.inhumation.defunt else 'Defunt'
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Autorisation_exhumation_{nom_defunt}_{timezone.now().strftime('%Y%m%d')}.pdf"'
    return response

@role_permission_required('generate_exhumation_pdf')
def pv_exhumation_pdf(request, demande_id):
    if not WEASYPRINT_AVAILABLE:
        messages.error(request, "Le système de génération PDF n'est pas disponible.")
        return redirect('portal_home')
        
    demande = get_object_or_404(DemandeExhumation, id=demande_id)
    context = {'demande': demande, 'parametres': ParametreCimetiere.objects.first(), 'date_generation': timezone.now(), 'site_name': 'Gestion Cimetière'}
    template = get_template('core/pdf/pv_exhumation_pdf.html')
    pdf_file = HTML(string=template.render(context), base_url=request.build_absolute_uri('/')).write_pdf()
    nom_defunt = demande.inhumation.defunt.nom.replace(' ', '_') if demande.inhumation and demande.inhumation.defunt else 'Defunt'
    date_str = demande.date_realisation.strftime('%Y%m%d') if demande.date_realisation else timezone.now().strftime('%Y%m%d')
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="PV_exhumation_{nom_defunt}_{date_str}.pdf"'
    return response

@role_permission_required('generate_statistical_report')
def rapport_statistique_pdf(request):
    if not WEASYPRINT_AVAILABLE:
        messages.error(request, "Le système de génération PDF n'est pas disponible.")
        return redirect('portal_home')
        
    debut_str = request.GET.get('debut')
    fin_str = request.GET.get('fin')
    if debut_str and fin_str:
        try:
            date_debut = datetime.strptime(debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(fin_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Format de date invalide.")
            return redirect('portal_home')
    else:
        today = timezone.now().date()
        date_debut = today.replace(day=1)
        _, last_day = calendar.monthrange(today.year, today.month)
        date_fin = today.replace(day=last_day)
    
    total_caveaux = Caveau.objects.count()
    caveaux_disponibles = Caveau.objects.filter(statut='DISPONIBLE').count()
    caveaux_occupes = Caveau.objects.filter(statut='OCCUPE').count()
    taux_occupation = round((caveaux_occupes / total_caveaux * 100), 1) if total_caveaux > 0 else 0
    concessions_actives = Concession.objects.filter(statut='ACTIVE').count()
    inhumations_periode = Inhumation.objects.filter(date_inhumation__range=(date_debut, date_fin)).count()
    exhumations_periode = DemandeExhumation.objects.filter(date_realisation__range=(date_debut, date_fin)).count()
    
    try:
        from apps.billing.models import Paiement
        revenus = Paiement.objects.filter(statut='VALIDE', date_paiement__range=(date_debut, date_fin)).aggregate(total=Sum('montant'))['total'] or 0
    except Exception:
        revenus = 0
    
    context = {
        'date_debut': date_debut, 'date_fin': date_fin, 'total_caveaux': total_caveaux,
        'caveaux_disponibles': caveaux_disponibles, 'caveaux_occupes': caveaux_occupes,
        'taux_occupation': taux_occupation, 'concessions_actives': concessions_actives,
        'inhumations_periode': inhumations_periode, 'exhumations_periode': exhumations_periode,
        'revenus_periode': revenus, 'parametres': ParametreCimetiere.objects.first(), 'date_generation': timezone.now(),
    }
    template = get_template('core/pdf/rapport_statistique_pdf.html')
    pdf_file = HTML(string=template.render(context), base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Rapport_Statistique_{date_debut.strftime('%Y%m')}_a_{date_fin.strftime('%Y%m%d')}.pdf"'
    return response


# ==============================================================================
# QR CODES CAVEAUX
# ==============================================================================
@role_permission_required('view_caveaux')
def qr_code_caveau(request, caveau_id):
    if not QR_AVAILABLE:
        messages.error(request, "La bibliothèque qrcode n'est pas installée.")
        return redirect('portal_home')
    caveau = get_object_or_404(Caveau, id=caveau_id)
    qr_url = request.build_absolute_uri(f'/cimetiere/caveau/{caveau.id}/qr-info/')
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(qr_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="image/png")
    response['Content-Disposition'] = f'inline; filename="qr_caveau_{caveau.code}.png"'
    return response

# Cette vue reste publique car elle est scannée par n'importe qui via le QR Code
def qr_info_caveau(request, caveau_id):
    caveau = get_object_or_404(Caveau, id=caveau_id)
    is_staff = request.user.is_staff if request.user.is_authenticated else False
    concession = None
    defunt = None
    if caveau.statut in ['OCCUPE', 'RESERVE']:
        concession = Concession.objects.filter(caveau=caveau).order_by('-date_debut').first()
        if concession and hasattr(concession, 'defunt') and concession.defunt:
            defunt = concession.defunt
    context = {
        'caveau': caveau, 'concession': concession, 'defunt': defunt,
        'is_staff': is_staff, 'parametres': ParametreCimetiere.objects.first()
    }
    return render(request, 'core/qr_info_caveau.html', context)


# ==============================================================================
# CONFIGURATION
# ==============================================================================
@role_permission_required('manage_settings')
def configurer_cimetiere(request):
    from .forms import ParametreCimetiereForm
    parametres = ParametreCimetiere.objects.first()
    if request.method == 'POST':
        form = ParametreCimetiereForm(request.POST, instance=parametres) if parametres else ParametreCimetiereForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '✓ Paramètres du cimetière enregistrés avec succès.')
            return redirect('configurer_cimetiere')
    else:
        form = ParametreCimetiereForm(instance=parametres) if parametres else ParametreCimetiereForm()
    
    total_zones = Zone.objects.count()
    total_caveaux = Caveau.objects.count()
    caveaux_disponibles = Caveau.objects.filter(statut='DISPONIBLE').count()
    capacite_theorique = 0
    if parametres and parametres.superficie_totale > 0:
        surface_caveau = (parametres.longueur_standard_caveau * parametres.largeur_standard_caveau) + (parametres.largeur_allee * parametres.longueur_standard_caveau)
        if surface_caveau > 0:
            capacite_theorique = int(parametres.superficie_totale / surface_caveau)
    
    context = {
        'title': '⚙️ Configuration du Cimetière', 'form': form, 'parametres': parametres,
        'total_zones': total_zones, 'total_caveaux': total_caveaux,
        'caveaux_disponibles': caveaux_disponibles, 'capacite_theorique': capacite_theorique,
    }
    return render(request, 'core/configurer_cimetiere.html', context)